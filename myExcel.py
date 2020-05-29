import openpyxl
import openpyxl.styles


class MyExcel:
    def __init__(self, _path):
        self.table_name_column = 1
        self.field_name_column = 2
        self.err_column = 3
        self.err_info_column = 4

        self.table_column = 1
        self.name_column = 2
        self.type_column = 3
        self.only_column = 4
        self.null_column = 5
        self.index_column = 6
        self.default_value_column = 7
        self.comment_column = 8
        self.exist_column = 9
        self.database_engine_column = 10
        self.default_charset_column = 11
        self.union_index_column = 12
        self._file_path = _path
        self._wxl = openpyxl.Workbook()
        self._sheet = self._wxl.create_sheet(title='err', index=0)

    def write(self, _err_map, _table_name):
        self._sheet.cell(row=1, column=self.table_column).value = '表名'
        self._sheet.cell(row=1, column=self.name_column).value = '字段名'
        self._sheet.cell(row=1, column=self.type_column).value = '类型'
        self._sheet.cell(row=1, column=self.only_column).value = '唯一'
        self._sheet.cell(row=1, column=self.null_column).value = '为空'
        self._sheet.cell(row=1, column=self.index_column).value = '索引'
        self._sheet.cell(row=1, column=self.default_value_column).value = '默认值'
        self._sheet.cell(row=1, column=self.comment_column).value = '注释'
        self._sheet.cell(row=1, column=self.exist_column).value = '缺失性'
        self._sheet.cell(row=1, column=self.database_engine_column).value = '数据库引擎'
        self._sheet.cell(row=1, column=self.default_charset_column).value = '数据库默认字符集'
        self._sheet.cell(row=1, column=self.union_index_column).value = '数据库联合索引'
        current_row = self._sheet.max_row + 1

        for _err_param in _err_map['params']:
            self._sheet.cell(row=current_row, column=self.table_column).value = _table_name
            self._sheet.cell(row=current_row, column=self.name_column).value = _err_param
            for _err in _err_map['params'][_err_param]:
                if _err == 'type_err':
                    self._sheet.cell(row=current_row, column=self.type_column).value = _err_map['params'][_err_param]['type_err']
                if _err == 'is_only_err':
                    self._sheet.cell(row=current_row, column=self.only_column).value = _err_map['params'][_err_param]['is_only_err']
                if _err == 'is_null_err':
                    self._sheet.cell(row=current_row, column=self.null_column).value = _err_map['params'][_err_param]['is_null_err']
                if _err == 'index_err':
                    self._sheet.cell(row=current_row, column=self.index_column).value = _err_map['params'][_err_param]['index_err']
                if _err == 'exist':
                    self._sheet.cell(row=current_row, column=self.exist_column).value = _err_map['params'][_err_param]['exist']
                if _err == 'default_err':
                    self._sheet.cell(row=current_row, column=self.default_value_column).value = _err_map['params'][_err_param]['default_err']
                if _err == 'comment_err':
                    self._sheet.cell(row=current_row, column=self.comment_column).value = _err_map['params'][_err_param]['comment_err']
            current_row += 1
        for table_err in _err_map['table']:
            self._sheet.cell(row=current_row, column=self.table_column).value = _table_name
            if table_err == 'database_engine_err':
                self._sheet.cell(row=current_row, column=self.database_engine_column).value = _err_map['table']['database_engine_err']
            if table_err == 'union_index_err':
                self._sheet.cell(row=current_row, column=self.union_index_column).value = _err_map['table']['union_index_err']
            if table_err == 'default_charset_err':
                self._sheet.cell(row=current_row, column=self.default_charset_column).value = _err_map['table']['default_charset_err']
            current_row += 1
        # for db_err in _err_map['db']:
        #     current_row += 1
        #     self._sheet.cell(row=current_row, column=1).value = '数据库缺失表'
        #     self._sheet.cell(row=current_row, column=2).value = '数据字典缺失表'
        #     if db_err == 'database_missing':
        #         self._sheet.cell(row=current_row, column=1).value = db_err + ''

    def merge_same_row(self, col_num, merge_type=1):
        start_row = 1
        past_value = ''
        max_row = self._sheet.max_row
        if max_row < 2:
            return
        if merge_type == 1:
            for i in range(2, max_row+1):
                current_value = self._sheet.cell(row=i, column=col_num).value
                if current_value != past_value:
                    self._sheet.merge_cells(start_row=start_row,end_row=i-1,start_column=col_num,end_column=col_num)
                    self._sheet.cell(row=start_row, column=col_num).alignment = openpyxl.styles.Alignment(vertical='center')
                    start_row = i
                    past_value = current_value
        if merge_type == 2:
            for i in range(2, max_row+1):
                current_value = self._sheet.cell(row=i, column=col_num).value
                if current_value == past_value and self._sheet.cell(row=i, column=col_num - 1).value == self._sheet.cell(row=i-1, column=col_num-1).value:
                    self._sheet.merge_cells(start_row=i-1,end_row=i,start_column=col_num,end_column=col_num)
                    self._sheet.cell(row=i-1, column=col_num).alignment = openpyxl.styles.Alignment(vertical='center')
                past_value = current_value

    def merge_null_col(self, start_col, end_col):
        pass

    def col_write(self, _err_map, _table_name):
        self._sheet.cell(row=1, column=self.table_name_column).value = '表名'
        self._sheet.cell(row=1, column=self.field_name_column).value = '字段名'
        self._sheet.cell(row=1, column=self.err_column).value = '错误类型'
        self._sheet.cell(row=1, column=self.err_info_column).value = '错误信息'
        self._sheet.cell(row=1, column=self.table_name_column).fill = openpyxl.styles.PatternFill("solid", fgColor="00FFFF")
        self._sheet.cell(row=1, column=self.field_name_column).fill = openpyxl.styles.PatternFill("solid", fgColor="00FFFF")
        self._sheet.cell(row=1, column=self.err_column).fill = openpyxl.styles.PatternFill("solid", fgColor="00FFFF")
        self._sheet.cell(row=1, column=self.err_info_column).fill = openpyxl.styles.PatternFill("solid", fgColor="00FFFF")

        current_row = self._sheet.max_row + 1
        for _err_param in _err_map['params']:
            self._sheet.cell(row=current_row, column=self.table_column).value = _table_name
            self._sheet.cell(row=current_row, column=self.name_column).value = _err_param
            for _err in _err_map['params'][_err_param]:
                if _err == 'type_err':
                    self._sheet.cell(row=current_row, column=self.table_name_column).value = _table_name
                    self._sheet.cell(row=current_row, column=self.field_name_column).value = _err_param
                    self._sheet.cell(row=current_row, column=self.err_column).value = u'类型错误'
                    self._sheet.cell(row=current_row, column=self.err_info_column).value = _err_map['params'][_err_param][
                        'type_err']
                    current_row += 1
                if _err == 'is_only_err':
                    self._sheet.cell(row=current_row, column=self.table_name_column).value = _table_name
                    self._sheet.cell(row=current_row, column=self.field_name_column).value = _err_param
                    self._sheet.cell(row=current_row, column=self.err_column).value = u'唯一性错误'
                    self._sheet.cell(row=current_row, column=self.err_info_column).value = _err_map['params'][_err_param][
                        'is_only_err']
                    current_row += 1
                if _err == 'is_null_err':
                    self._sheet.cell(row=current_row, column=self.table_name_column).value = _table_name
                    self._sheet.cell(row=current_row, column=self.field_name_column).value = _err_param
                    self._sheet.cell(row=current_row, column=self.err_column).value = u'为空错误'
                    self._sheet.cell(row=current_row, column=self.err_info_column).value = _err_map['params'][_err_param][
                        'is_null_err']
                    current_row += 1
                if _err == 'index_err':
                    self._sheet.cell(row=current_row, column=self.table_name_column).value = _table_name
                    self._sheet.cell(row=current_row, column=self.field_name_column).value = _err_param
                    self._sheet.cell(row=current_row, column=self.err_column).value = u'索引错误'
                    self._sheet.cell(row=current_row, column=self.err_info_column).value = _err_map['params'][_err_param][
                        'index_err']
                    current_row += 1
                if _err == 'exist':
                    self._sheet.cell(row=current_row, column=self.table_name_column).value = _table_name
                    self._sheet.cell(row=current_row, column=self.field_name_column).value = _err_param
                    self._sheet.cell(row=current_row, column=self.err_column).value = u'字段缺失'
                    self._sheet.cell(row=current_row, column=self.err_info_column).value = _err_map['params'][_err_param][
                        'exist']
                    current_row += 1
                if _err == 'default_err':
                    self._sheet.cell(row=current_row, column=self.table_name_column).value = _table_name
                    self._sheet.cell(row=current_row, column=self.field_name_column).value = _err_param
                    self._sheet.cell(row=current_row, column=self.err_column).value = u'默认值错误'
                    self._sheet.cell(row=current_row, column=self.err_info_column).value = _err_map['params'][_err_param]['default_err']
                    current_row += 1
                if _err == 'comment_err':
                    self._sheet.cell(row=current_row, column=self.table_name_column).value = _table_name
                    self._sheet.cell(row=current_row, column=self.field_name_column).value = _err_param
                    self._sheet.cell(row=current_row, column=self.err_column).value = u'注释错误'
                    self._sheet.cell(row=current_row, column=self.err_info_column).value = _err_map['params'][_err_param]['comment_err']
                    current_row += 1
        for table_err in _err_map['table']:
            if table_err == 'database_engine_err':
                self._sheet.cell(row=current_row, column=self.table_name_column).value = _table_name
                self._sheet.cell(row=current_row, column=self.err_column).value = u'数据表存储引擎错误'
                self._sheet.cell(row=current_row, column=self.err_info_column).value = _err_map['table'][
                    'database_engine_err']
                current_row += 1
            if table_err == 'union_index_err':
                self._sheet.cell(row=current_row, column=self.table_name_column).value = _table_name
                self._sheet.cell(row=current_row, column=self.err_column).value = u'联合索引错误'
                self._sheet.cell(row=current_row, column=self.err_info_column).value = _err_map['table'][
                    'union_index_err']
                current_row += 1
            if table_err == 'default_charset_err':
                self._sheet.cell(row=current_row, column=self.table_name_column).value = _table_name
                self._sheet.cell(row=current_row, column=self.err_column).value = u'默认字符集错误'
                self._sheet.cell(row=current_row, column=self.err_info_column).value = _err_map['table'][
                    'default_charset_err']
                current_row += 1
        self.merge_same_row(self.table_name_column)
        self.merge_same_row(self.field_name_column, 2)

    def save(self):
        self._wxl.save(self._file_path)
        self._wxl.close()
